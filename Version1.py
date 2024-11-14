from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

archivo_excel = 'registro.xlsx'

# Verificar si el archivo existe
if not os.path.exists(archivo_excel):
    wb = Workbook()
    ws = wb.active
    ws.append(["Razon Social", "RUT", "Direccion", "Tipo Comprobante", "Fecha", "Codigo Cuenta", "Detalle", "Monto"])
    wb.save(archivo_excel)

while True:
    razon_social = input("Ingrese la razon social: ")
    print("La razon social es:", razon_social)

    rut = input("Ingrese el rut: ")
    print("El rut es:", rut)

    direccion = input("Ingrese la direccion: ")
    print("La direccion es:", direccion)

    print("Elija el tipo de comprobante (1: Ingreso 2: Egreso)")
    tipo_comprobante = input("Ingrese el tipo de comprobante: ")
    while tipo_comprobante not in ["1", "2"]:
        print("Error, ingrese un tipo de comprobante valido")
        tipo_comprobante = input("Ingrese el tipo de comprobante: ")

    fecha_valida = False
    while not fecha_valida:
        fecha = input("Ingrese la fecha (DD/MM/YYYY): ")
        try:
            datetime.strptime(fecha, "%d/%m/%Y")
            fecha_valida = True
        except ValueError:
            print("Error, ingrese una fecha valida en el formato DD/MM/YYYY")

    print("La fecha es:", fecha)
    codigo_cuenta = input("Ingrese el codigo de cuenta: ")
    print("El codigo de cuenta es:", codigo_cuenta)

    detalle = input("Ingrese el detalle (Compra/Venta): ")
    while detalle not in ["Compra", "Venta"]:
        print("Error, ingrese un detalle valido")
        detalle = input("Ingrese el detalle (Compra/Venta): ")

    if tipo_comprobante == "1":
        debe = float(input("Ingrese el monto: "))
        print("El monto es:", debe)
        monto = debe
        tipo = "Debe"
    elif tipo_comprobante == "2":
        haber = float(input("Ingrese el monto: "))
        print("El monto es:", haber)
        monto = haber
        tipo = "Haber"

    try:
        # Cargar el archivo Excel
        wb = load_workbook(archivo_excel)
        ws = wb.active

        # Agregar los datos
        ws.append([razon_social, rut, direccion, tipo, fecha, codigo_cuenta, detalle, monto])

        # Guardar el archivo
        wb.save(archivo_excel)
        print("Comprobante agregado exitosamente.")
    except Exception as e:
        print(f"Ocurrió un error al guardar el comprobante: {e}")

    continuar = input("¿Desea agregar otro comprobante? (s/n): ").lower()
    if continuar != 's':
        break